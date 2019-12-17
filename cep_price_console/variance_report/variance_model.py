from cep_price_console.utils.gui_utils import OrderableListItemCheckbutton, OrderableListItemEntryCheckbutton
from cep_price_console.utils.log_utils import CustomAdapter, debug
from copy import copy
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.sql.expression import and_, or_
from tkinter import messagebox
import logging
import pandas as pd
import pathlib
import numpy as np
import re
from collections import OrderedDict

class SalesVarianceModel(object):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)

    @debug(lvl=logging.NOTSET, prefix='')
    def __init__(self, view):
        self.view = view

        self.available_comparisons = {
            "var": Comparison(header="Variance", abbrev="Var."),
            "pcnt_var": Comparison(header="Percentage Variance", abbrev="% Var.")
        }

        self.selected_col_dict = {
            "ProdNumInv": ProdNumInv(),
            "CustPartNum": CustPartNum(),
            "DisplayUOM": DisplayUOM(),
            "DescFull": DescFull(),
            "PrimVendPartNum": PrimVendPartNum()
        }

        self.available_col_dict = {
            # "": ProdNumMain(),
            "C1Cost": C1Cost(),
            "L1Price": L1Price(),
            "MfgNum": MfgNum(),
            "PriceGroup": PriceGroup(show_check=True, checkbox_label="Group On"),
            "PriceUOM": PriceUOM(),
            "PrimVend": PrimVend(show_check=True, checkbox_label="Group On"),
            "ProdLine": ProdLine(show_check=True, checkbox_label="Group On"),
            "PurchUOM": PurchUOM(),
            "Status": Status(),
            "UOMFactorDesc": UOMFactorDesc(),
            "UPC": UPC(),
            "Commission": Commission(),
            "Cost": Cost(),
            "GLCost": GLCost(),
            "GLProfit": GLProfit(),
            "OriginalQuantity": OriginalQuantity(),
            "Profit": Profit(),
            "QuantityShipped": QuantityShipped(),
            "Revenue": Revenue(),
            "SalesmanCost": SalesmanCost(),
            "SalesmanProfit": SalesmanProfit(),
        }

        from cep_price_console.db_management.server_utils import mysql_session_maker
        self.session = mysql_session_maker()
        self.date_item_combo = None
        self.__fullpath = None
        self.__excelwriter = None
        self.__col_int = 100
        self.col_dict = OrderedDict()
        self.include_summary = True
        self.compare_summary = True
        self.include_individual = True
        self.compare_individual = False
        self.customer_list = None
        self.formatting_index_df = None
        self.master_df = None

    @debug(lvl=logging.DEBUG, prefix='')
    def set_formatting_index_df(self):
        formatting_df = pd.DataFrame(data={
            "Cust_Num": ["0000_Formatting"],
            self.col_prod_num_inv.entry_var.get(): ["0000_Formatting"]
        })
        formatting_df.set_index("Cust_Num", inplace=True, drop=True)
        formatting_df.set_index(self.col_prod_num_inv.entry_var.get(), inplace=True, append=True, drop=True)
        self.__class__.logger.log(logging.NOTSET, formatting_df.to_string())
        self.formatting_index_df = formatting_df

    @debug(lvl=logging.NOTSET, prefix='')
    def set_customer_list(self):
        self.customer_list = list(set(self.master_df.index.get_level_values('Cust_Num').tolist()))
        self.set_formatting_index_df()

    @debug(lvl=logging.NOTSET, prefix='')
    def close(self):
        return self.session.close()

    # region tab_text
    @property
    @debug(lvl=logging.NOTSET, prefix='')
    def fullpath(self):
        return self.__fullpath

    @property
    @debug(lvl=logging.NOTSET, prefix='')
    def excelwriter(self):
        return self.__excelwriter

    @debug(lvl=logging.NOTSET, prefix='')
    def set_fullpath(self, value):
        fullpath = pathlib.Path(value)
        if len(fullpath.suffixes) == 0:
            fullpath = fullpath.with_suffix(".xlsx")
        elif len(fullpath.suffixes) > 1:
            messagebox.askokcancel(
                "Too Many Suffixes",
                "Only one suffix allowed: .xlsx",
                parent=self.view
            )
            return False
        if fullpath.suffix == ".xlsx":
            self.__class__.logger.log(
                logging.DEBUG,
                "set_fullpath called. Filepath valid. Extension valid. Filepath: {0}".format(value))
            try:
                fullpath.parent.resolve(strict=True)
            except FileNotFoundError:
                messagebox.askokcancel(
                    "Folder Not Found",
                    "Folder not found: {0}".format(value),
                    parent=self.view
                )
                return False
            else:
                self.__fullpath = fullpath
                self.__excelwriter = pd.ExcelWriter(
                    self.fullpath,
                    engine="openpyxl",
                    date_format="MM/DD/YYYY",
                    datetime_format="MM/DD/YYYY",
                    mode=self.excelwriter_mode()
                )
                # noinspection PyPep8
                try:
                    self.styles()
                except:
                    pass
                return True
        else:
            self.__class__.logger.log(
                logging.DEBUG,
                "set_fullpath: Invalid file extension messagebox called")
            messagebox.askokcancel(
                "Invalid Extension",
                "Invalid file extension. Supported extensions include: .xlsx",
                parent=self.view
            )
            return False

    # endregion
    @property
    @debug(lvl=logging.NOTSET, prefix='')
    def col_int(self):
        self.__col_int += 1
        return self.__col_int

    @debug(lvl=logging.DEBUG, prefix="")
    def get_col_label(self, label_obj, format_obj, col_int=None):
        if col_int is None:
            label = "({:03}) {} {}".format(self.col_int, str(format_obj.__class__.__name__), str(label_obj))
        else:
            label = "({:03}) {} {}".format(col_int, str(format_obj.__class__.__name__), str(label_obj))

        self.col_dict[label] = {
            "label_obj": label_obj,
            "format_obj": format_obj,
            "final_label": label_obj,
        }
        # self.__class__.logger.log(logging.DEBUG, "{}: {}".format(label, str(type(format_obj))))
        return label

    @property
    @debug(lvl=logging.DEBUG, prefix="")
    def col_list(self):
        return self.view.report_options_frame.column_selection.item_dict.values()

    @property
    @debug(lvl=logging.DEBUG, prefix="")
    def col_list_metric(self):
        return [col_obj for col_obj in self.col_list if col_obj.metric]

    @property
    @debug(lvl=logging.DEBUG, prefix="")
    def col_list_nonmetric(self):
        return [col_obj for col_obj in self.col_list if not col_obj.metric]

    @property
    @debug(lvl=logging.DEBUG, prefix="")
    def col_list_grouping(self):
        return [col_obj for col_obj in self.col_list_nonmetric if 'selected' in col_obj.sel_check_btn.state()]

    @property
    @debug(lvl=logging.DEBUG, prefix="")
    def col_list_header(self):
        return [col_obj for col_obj in self.col_list_nonmetric if 'selected' not in col_obj.sel_check_btn.state()]

    @property
    @debug(lvl=logging.NOTSET, prefix="")
    def comparisons(self):
        return [comparison for comparison in self.view.report_options_frame.comparison_frame.item_dict.values()
                if 'selected' in comparison.sel_check_btn.state()]

    @debug(lvl=logging.NOTSET, prefix="")
    def excelwriter_mode(self):
        if self.fullpath.is_file():
            return 'a'
        else:
            return 'w'

    @debug(lvl=logging.NOTSET, prefix="")
    def get_col(self, col_label):
        for col in self.col_list:
            if col.label == col_label:
                return col
        messagebox.askokcancel(
            "Critical Error",
            "{} column not found. Please restart application.".format(col_label),
            parent=self.view
        )
        self.view.master.quit()

    @property
    @debug(lvl=logging.NOTSET, prefix="")
    def col_prod_num_inv(self):
        return self.get_col("Prod_Num")

    @property
    @debug(lvl=logging.NOTSET, prefix="")
    def col_quantity_shipped(self):
        return self.get_col("Quantity_Shipped")

    @debug(lvl=logging.DEBUG, prefix='')
    def create_wb(self, prod_list, cust_shipto_list, cust_list):
        master_df = self.get_master_df(prod_list=prod_list, cust_shipto_list=cust_shipto_list, cust_list=cust_list)
        if master_df is not None:
            cust_list = list(set(master_df.index.get_level_values('Cust_Num').tolist()))
            cust_list.remove("0000_Formatting")
            self.__class__.logger.log(logging.DEBUG, "cust_list = {}".format(cust_list))
            for cust_num in sorted(cust_list):
                cust_df = master_df.loc[(["0000_Formatting", cust_num])]
                cust_df.reset_index(level="Cust_Num", drop=True, inplace=True)
                self.__class__.logger.log(logging.NOTSET,
                                          "Single-Customer Dataframe: \n{}".format(cust_df.to_string()))
                self.create_ws(cust_num=cust_num, cust_df=cust_df)
            self.excelwriter.save()
        else:
            pass

    @debug(lvl=logging.DEBUG, prefix='')
    def get_master_df(self, prod_list, cust_shipto_list, cust_list):
        master_index = self.get_master_index(cust_list=cust_list, prod_list=prod_list)
        if len(self.master_df) != 0:
            return_df = None
            for date_list in self.date_item_combo.get_comparison_range_list(
                    include_summary=self.include_summary,
                    compare_summary=self.compare_summary,
                    include_individual=self.include_individual,
                    compare_individual=self.compare_individual):
                df_combo = self.master_range_df(master_index=master_index, date_list=date_list)
                if return_df is None:
                    return_df = df_combo
                else:
                    return_df = return_df.join(df_combo, how='outer').fillna(0)
            return_df.sort_index(axis=1, level="metric", inplace=True)
            self.__class__.logger.log(logging.NOTSET,
                                      "Master Dataframe Pre-Grouping: \n{}".format(return_df.to_string()))

            return_df = return_df.join(
                self.get_grouping_df(list(set(
                    master_index.index.get_level_values(self.col_prod_num_inv.entry_var.get()).tolist()))),
                how="outer"
            )

            self.__class__.logger.log(logging.NOTSET,
                                      "Master Dataframe Post-Grouping: \n{}".format(return_df.to_string()))
            header_dict = {col_obj.entry_var.get(): "({:0>3}) {}".format(count, col_obj.entry_var.get())
                           for count, col_obj in enumerate(self.col_list_header, 1)}
            for col_label, new_col_label in header_dict.items():
                self.__class__.logger.log(logging.DEBUG,
                                          "Old Label: {} New Label: {}".format(col_label, new_col_label))
            return_df.rename(columns=header_dict, inplace=True)
            return_df.sort_index(axis=1, level="metric", inplace=True)

            self.__class__.logger.log(logging.NOTSET,
                                      "Master Dataframe Non-Metric Columns in place: \n{}".format(return_df.to_string()))

            grouping_list = [col_obj.entry_var.get() for col_obj in self.col_list_grouping]
            for group in grouping_list:
                # return_df.set_index([("", group)], append=True, inplace=True)
                return_df.set_index(
                    pd.Index(return_df[("", group)], name=group),
                    append=True, inplace=True)
                return_df.drop(columns=[("", group)], inplace=True)

            return_df = return_df.reorder_levels(
                ["Cust_Num", *grouping_list, self.col_prod_num_inv.entry_var.get()],
                axis=0
            ).sort_index(axis=0, level="Cust_Num")

            self.__class__.logger.log(logging.NOTSET,
                                      "Master Dataframe Pre-Customer Split: \n{}".format(return_df.to_string()))
            return return_df
        else:
            return None

    @debug(lvl=logging.DEBUG, prefix='')
    def get_master_index(self, cust_list, prod_list):
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        index_query = self.session.query(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.label("Cust_Num"),
            self.col_prod_num_inv.labeled_col()
        ).join(
            ddi_data_warehouse.invoice_detail,
            ddi_data_warehouse.invoice_header.c.INVOICECODE ==
            ddi_data_warehouse.invoice_detail.c.INVOICECODE
        ).filter(and_(
            self.col_quantity_shipped.column != 0,
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.in_(cust_list),
            ProdNumInv.column().in_(prod_list),
            or_(*self.get_date_filter_list()))
        ).group_by(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE,
            ProdNumInv.column()
        ).order_by(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.asc(),
            ProdNumInv.column().asc()
        )

        # from cep_price_console.db_management.server_utils import verbose_query
        # verbose_query(index_query)

        return_df = pd.read_sql(
            index_query.statement,
            self.session.bind,
            index_col=["Cust_Num", self.col_prod_num_inv.entry_var.get()]
        )

        self.__class__.logger.log(logging.NOTSET, "get_master_index:\n" + return_df.to_string())

        self.master_df = return_df
        self.set_customer_list()
        return return_df

    @debug(lvl=logging.DEBUG, prefix='')
    def get_date_filter_list(self):
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        date_filter_list = []
        for date_list in self.date_item_combo.get_comparison_range_list(
                include_summary=True,
                compare_summary=False,
                include_individual=False,
                compare_individual=False):
            for date_obj in date_list:
                date_filter_list.append(
                    ddi_data_warehouse.invoice_header.c.INVOICEDATE.between(
                        str(date_obj.start_date.strftime("%Y-%m-%d %H:%M:%S")),
                        str(date_obj.end_date.strftime("%Y-%m-%d %H:%M:%S"))))
        return date_filter_list

    @debug(lvl=logging.DEBUG, prefix='')
    def master_range_df(self, master_index, date_list):
        date_list = sorted(date_list, key=lambda x: x.start_date)
        return_df = None
        for date_item in date_list:
            sub_df = self.master_all_metric_df(master_index=master_index, date_item=date_item)
            if return_df is None:
                return_df = sub_df
            else:
                return_df = return_df.join(sub_df, how='outer').fillna(0)

        self.__class__.logger.log(logging.NOTSET, "master_range_df pre-reorder:\n" + return_df.to_string())
        return_df = return_df.reorder_levels(["metric", "comparison"], axis=1).sort_index(axis=1, level="metric")
        self.__class__.logger.log(logging.NOTSET, "master_range_df post-reorder:\n" + return_df.to_string())

        if date_list[0].date_type in ("month", "quarter"):
            bln_val = self.compare_individual
        elif date_list[0].date_type in ("month_summary", "quarter_summary", "year"):
            bln_val = self.compare_summary
        else:
            bln_val = None
            messagebox.askokcancel(
                "Critical Error",
                "Unknown Date Item Value: {}. Please restart the application.".format(date_list[0].date_type),
                parent=self.view
            )
            self.view.master.quit()

        if bln_val:
            comp_list = [comp.header for comp in self.comparisons]

            date_label = "({:03}) {}".format(self.__col_int, str(date_list[-1]))
            if "Variance" in comp_list:
                comp_list.remove("Variance")
                col_label = "({:03}) {} Var.".format(self.col_int, str(date_list[-1]))
                for metric_obj in self.col_list_metric:
                    # noinspection PyTypeChecker
                    return_df[metric_obj.entry_var.get(), col_label] = \
                        return_df.drop("0000_Formatting", axis=0, level=1).diff(
                            axis=1
                        ).loc[
                        :, (slice(metric_obj.entry_var.get(), metric_obj.entry_var.get()),
                            slice(date_label, date_label))
                        ].fillna(0)
                    return_df.sort_index(axis=1, level="metric", inplace=True)
                    return_df.loc[(slice(None), slice("0000_Formatting")),
                                          (slice(metric_obj.entry_var.get(), metric_obj.entry_var.get()),
                                           slice(col_label, col_label))] = metric_obj.__class__.cell_format

            if "Percentage Variance" in comp_list:
                comp_list.remove("Percentage Variance")
                col_label = "({:03}) {} % Var.".format(self.col_int, str(date_list[-1]))
                for metric_obj in self.col_list_metric:
                    # noinspection PyTypeChecker
                    return_df[metric_obj.entry_var.get(), col_label] = \
                        return_df.drop("0000_Formatting", axis=0, level=1).pct_change(
                            axis=1
                        ).loc[
                        :, (slice(metric_obj.entry_var.get(), metric_obj.entry_var.get()),
                            slice(date_label, date_label))
                        ].fillna(0).replace(
                            [np.inf], 1
                        )
                    return_df.sort_index(axis=1, level="metric", inplace=True)
                    return_df.loc[(slice(None), slice("0000_Formatting")),
                                          (slice(metric_obj.entry_var.get(), metric_obj.entry_var.get()),
                                           slice(col_label, col_label))] = "InteriorPercentCell"
            if comp_list:
                messagebox.askokcancel(
                    "Critical Error",
                    "Unknown Comparison Value: {}. Please restart the application.".format(comp_list),
                    parent=self.view
                )
                self.view.master.quit()
            self.__class__.logger.log(logging.NOTSET, "master_range_df post-comparison:\n" + return_df.to_string())
        return return_df

    @debug(lvl=logging.DEBUG, prefix='')
    def master_all_metric_df(self, master_index, date_item):
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        formatting_df = self.formatting_index_df.copy()

        metric_sum_list = []
        for metric in self.col_list_metric:
            metric_sum_list.append(func.sum(metric.column()).label(metric.entry_var.get()))
            formatting_df[metric.entry_var.get()] = metric.__class__.cell_format

        self.__class__.logger.log(logging.NOTSET, "Formatting DF: \n{}".format(formatting_df.to_string()))

        query = self.session.query(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.label("Cust_Num"),
            self.col_prod_num_inv.labeled_col(),
            *metric_sum_list
        ).outerjoin(
            ddi_data_warehouse.invoice_detail,
            ddi_data_warehouse.invoice_header.c.INVOICECODE ==
            ddi_data_warehouse.invoice_detail.c.INVOICECODE
        ).filter(and_(
            self.col_quantity_shipped.column != 0,
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.in_(self.customer_list),
            ProdNumInv.column().in_(
                list(set(master_index.index.get_level_values(self.col_prod_num_inv.entry_var.get()).tolist()))),
            ddi_data_warehouse.invoice_header.c.INVOICEDATE.between(
                str(date_item.start_date.strftime("%Y-%m-%d %H:%M:%S")),
                str(date_item.end_date.strftime("%Y-%m-%d %H:%M:%S"))))
        ).group_by(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE,
            ProdNumInv.column()
        ).order_by(
            ddi_data_warehouse.invoice_header.c.CUSTOMERCODE.asc(),
            ProdNumInv.column().asc()
        )

        # from cep_price_console.db_management.server_utils import verbose_query
        # verbose_query(query)
        return_df = pd.read_sql(
            query.statement,
            self.session.bind,
            index_col=["Cust_Num", self.col_prod_num_inv.entry_var.get()]
        )
        self.__class__.logger.log(logging.NOTSET, "Pre MultiIndex: \n{}".format(return_df.to_string()))

        return_df = return_df.append(formatting_df)
        self.__class__.logger.log(logging.NOTSET, "Append Format Pre-Sort: \n{}".format(return_df.to_string()))
        return_df.sort_index(axis=0, level="Cust_Num", inplace=True)
        self.__class__.logger.log(logging.NOTSET, "Append Format Post-Sort: \n{}".format(return_df.to_string()))

        return_df.columns = pd.MultiIndex.from_product(
            [["({:03}) {}".format(self.col_int, str(date_item))], return_df.columns],
            names=["comparison", "metric"]
        )

        self.__class__.logger.log(logging.NOTSET, "Post Multiindex From Product: \n{}".format(return_df.to_string()))
        return return_df

    @debug(lvl=logging.NOTSET, prefix='')
    def get_grouping_df(self, prod_list):
        col_list = [col_obj.labeled_col() for col_obj in self.col_list_nonmetric]
        index_query = self.session.query(
            *col_list
        )
        table_set = set()
        for col in self.col_list_nonmetric:
            if col.table() not in table_set:
                table_set.add(col.table())
                index_query = col.apply_join(index_query)

        index_query = index_query.filter(
            ProdNumInv.column().in_(prod_list)
        ).order_by(*self.get_ordering_list()).group_by(
            ProdNumInv.column()
        )

        # from cep_price_console.db_management.server_utils import verbose_query
        # verbose_query(index_query)

        return_df = pd.read_sql(
            index_query.statement,
            self.session.bind,
            index_col=[self.col_prod_num_inv.entry_var.get()]
        )
        self.__class__.logger.log(logging.NOTSET, return_df.to_string())

        return_df.columns = pd.MultiIndex.from_product(
            [[""], return_df.columns],
            names=["comparison", "metric"]
        )
        self.__class__.logger.log(logging.NOTSET, return_df.to_string())

        return return_df

    @debug(lvl=logging.DEBUG, prefix='')
    def get_ordering_list(self):
        col_list = [col_obj.labeled_col() for col_obj in self.col_list_grouping]
        ordering_list = []
        for col in reversed(col_list):
            self.__class__.logger.log(logging.DEBUG, "Column Name: {0}".format(col.name))
            ordering_list.append(col.asc())
        self.__class__.logger.log(logging.DEBUG, "Ordering List: {0}".format(ordering_list))
        return ordering_list

    @debug(lvl=logging.DEBUG, prefix="")
    def create_ws(self, cust_num, cust_df):
        first_col = 1
        first_row = 1
        wb = self.excelwriter.book

        sheet_count = 0
        cust_num_sheet_name = None
        while cust_num_sheet_name is None:
            trial = "{}_{:02}".format(cust_num, sheet_count)
            if trial in wb.sheetnames:
                sheet_count += 1
            else:
                cust_num_sheet_name = trial

        df = cust_df
        df.to_excel(
            excel_writer=self.excelwriter,
            # float_format="%.2f",
            startrow=0,
            startcol=0,
            merge_cells=True,
            freeze_panes=(df.columns.nlevels, df.index.nlevels),
            na_rep="None",
            sheet_name=cust_num_sheet_name
        )

        ws = wb[cust_num_sheet_name]
        for index_col in range(first_col, df.index.nlevels + 1):
            ws.cell(
                column=index_col, row=df.columns.nlevels
            ).value = ws.cell(
                column=index_col, row=(df.columns.nlevels + 1)
            ).value

        # Format worksheet
        page_setup = ws.page_setup
        page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        page_setup.paperSize = ws.PAPERSIZE_LETTER

        ws.page_margins.left = .25
        ws.page_margins.right = .25
        ws.page_margins.top = .75
        ws.page_margins.bottom = .75
        ws.page_margins.header = .3
        ws.page_margins.footer = .3

        print_options = ws.print_options
        print_options.horizontalCentered = True
        print_options.verticalCentered = False

        hdr = ws.oddFooter
        hdr.left.text = "&[Path]&[File]"
        hdr.right.text = "Page &[Page] of &N"

        ws.sheet_properties.tabColor = "1072BA"

        # print("Rows: {}".format(len(df.index)))
        print("Index Rows: {}".format(df.index.nlevels))

        # print("Columns: {}".format(len(df.columns)))
        # print("Index Columns: {}".format(df.columns.nlevels))

        ws.print_title_cols = "{first_col}:{final_col}".format(
            first_col=get_column_letter(first_col),
            final_col=get_column_letter(df.index.nlevels))
        ws.print_title_rows = "{first_row}:{final_row}".format(
            first_row="1",
            final_row=df.columns.nlevels)
        ws.print_area = "{first_col}{first_row}:{final_col}{final_row}".format(
            first_col=get_column_letter(1),
            first_row="1",
            final_col=get_column_letter(len(df.columns) + df.index.nlevels),
            final_row=str(len(df.index) + df.columns.nlevels + 1))

        for merge in sorted(ws.merged_cells):
            self.__class__.logger.log(logging.NOTSET, "Merged Cell: {}".format(merge))

        for col_id, col in enumerate(ws.iter_cols(
                min_row=4, max_row=4), 1):
            for header_cell in col:
                self.__class__.logger.log(logging.NOTSET, "Format: {}".format(header_cell.value))
                for row_id, row in enumerate(ws.iter_rows(min_row=5, min_col=col_id, max_col=col_id), 1):
                    for cell in row:
                        self.__class__.logger.log(
                            logging.NOTSET, "Format: {} Cell Value: {}".format(header_cell.value, cell.value))
                        if header_cell.value not in ("None", "0000_Formatting"):
                            cell.style = header_cell.value
                        else:
                            cell.style = "InteriorTextCell"
                header_cell.value = ""

        for row_id, row in enumerate(ws.iter_rows(), 1):
            for col_id, cell in enumerate(row, 1):
                self.__class__.logger.log(logging.NOTSET,
                                          "Row: {} Col: {} Value: {}".format(row_id, col_id, cell.value))
                if row_id in range(first_row, df.columns.nlevels):
                    cell.style = 'ColHdrMain'
                elif row_id == df.columns.nlevels:
                    cell.style = 'ColHdrBottom'
                    if "% Var." in cell.value:
                        cell.value = "% Var."
                    elif "Var." in cell.value:
                        cell.value = "Var."
                    cell.value = re.sub(r'\(.*\)', '', cell.value)
                elif col_id in range(first_col, df.index.nlevels):
                    cell.style = 'IndexMain'
                elif col_id == df.index.nlevels:
                    cell.style = 'IndexEnd'

    @debug(lvl=logging.NOTSET, prefix="")
    def styles(self):
        wb = self.excelwriter.book

        col_hdr_main = NamedStyle('ColHdrMain')
        col_hdr_main.font = Font(bold=True, size=12)
        col_hdr_main.alignment = Alignment(horizontal='center', wrap_text=True)
        col_hdr_main.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        wb.add_named_style(col_hdr_main)

        col_hdr_bottom = copy(col_hdr_main)
        col_hdr_bottom.name = "ColHdrBottom"
        col_hdr_bottom.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='medium')
        )
        wb.add_named_style(col_hdr_bottom)

        index_main = NamedStyle('IndexMain')
        index_main.font = Font(bold=True, size=12)
        index_main.alignment = Alignment(horizontal='left', vertical="top", wrap_text=False)
        index_main.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        wb.add_named_style(index_main)

        index_end = copy(index_main)
        index_end.name = "IndexEnd"
        index_end.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='medium'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        wb.add_named_style(index_end)

        interior_cell = NamedStyle('InteriorCell')
        interior_cell.font = Font(size=11)
        interior_cell.alignment = Alignment(horizontal='right', wrap_text=False)
        interior_cell.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        wb.add_named_style(interior_cell)

        interior_int_cell = copy(interior_cell)
        interior_int_cell.name = "InteriorIntCell"
        interior_int_cell.number_format = "#,##0;[Red]#,##0;-;"
        wb.add_named_style(interior_int_cell)

        interior_currency_cell = copy(interior_cell)
        interior_currency_cell.name = "InteriorCurrencyCell"
        interior_currency_cell.number_format = "#,##0.00_);[Red](#,##0.00);-;"
        wb.add_named_style(interior_currency_cell)

        interior_percent_cell = copy(interior_cell)
        interior_percent_cell.name = "InteriorPercentCell"
        interior_percent_cell.number_format = "0%;[Red]0%;-;"
        wb.add_named_style(interior_percent_cell)

        interior_text_cell = copy(interior_cell)
        interior_text_cell.name = "InteriorTextCell"
        interior_text_cell.number_format = "General"
        wb.add_named_style(interior_text_cell)


class Comparison(OrderableListItemCheckbutton):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)

    @debug(lvl=logging.NOTSET, prefix='')
    def __init__(self, header, abbrev):
        self.header = header
        self.abbrev = abbrev
        super().__init__(title=self.header)

    @debug(lvl=logging.NOTSET, prefix="")
    def grid_init(self, key, list_frame, parent, **kwargs):
        super().grid_init(key, list_frame, parent, **kwargs)
        self.sel_check_btn.state(['selected'])

    def __str__(self):
        return self.abbrev

    @debug(lvl=logging.NOTSET, prefix='')
    def get_comparison_series(self, second_series, first_series):
        if self.header == "Variance":
            return second_series - first_series
        elif self.header == "Percentage Variance":
            return (second_series - first_series) / 100


class ColumnObject(OrderableListItemEntryCheckbutton):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = None
    header = None
    abbrev = None
    metric = False
    cell_format = "InteriorTextCell"

    @debug(lvl=logging.NOTSET, prefix='')
    def __init__(self, show_check=False, checkbox_label=""):
        self.label = self.__class__.label
        self.header = self.__class__.header
        self.abbrev = self.__class__.abbrev
        self.metric = self.__class__.metric
        super().__init__(title=self.header,
                         dflt_entry=self.abbrev,
                         show_check=show_check,
                         checkbox_label=checkbox_label)

    @debug(lvl=logging.NOTSET, prefix='')
    def labeled_col(self):
        return self.column().label(self.entry_var.get())

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def table(self):
        raise NotImplementedError

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        raise NotImplementedError

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def apply_join(query):
        raise NotImplementedError

    @debug(lvl=logging.NOTSET, prefix='')
    def __str__(self):
        return self.abbrev


# region InvoiceDetailColumn  ##################################################################################
class InvoiceDetailColumn(ColumnObject):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = None
    header = None
    abbrev = None
    metric = False
    cell_format = "InteriorTextCell"

    @staticmethod
    def apply_join(query):
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        return query.outerjoin(
            ddi_data_warehouse.invoice_header,
            ddi_data_warehouse.invoice_header.c.INVOICECODE ==
            ddi_data_warehouse.invoice_detail.c.INVOICECODE
        )

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def table():
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        return ddi_data_warehouse.invoice_detail


class ProdNumInv(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Prod_Num"
    header = "Product Num"
    abbrev = "Prod. Num"
    metric = False
    cell_format = "InteriorTextCell"

    @debug(lvl=logging.NOTSET, prefix="")
    def grid_init(self, key, list_frame, parent, **kwargs):
        super().grid_init(key, list_frame, parent, **kwargs)
        # self.sel_check_btn.state(['selected', 'disabled'])
        self.delete_item.grid_forget()

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.PRODUCTCODE


class QuantityShipped(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Quantity_Shipped"
    header = "Quantity Shipped"
    abbrev = "Qty. Shipped"
    metric = True
    cell_format = "InteriorIntCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.QUANTITYSHIPPED


class OriginalQuantity(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Original_Quantity"
    header = "Original Quantity"
    abbrev = "Orig. Qty."
    metric = True
    cell_format = "InteriorIntCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.ORIGINALQUANTITY


class Cost(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Cost"
    header = "Cost"
    abbrev = "Cost"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDCOST


class GLCost(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "GL_Cost"
    header = "GL Cost"
    abbrev = "GL Cost"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDGLCOST


class SalesmanCost(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Salesman_Cost"
    header = "Salesman Cost"
    abbrev = "Salesman Cost"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDSLMNCOST


class Revenue(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Revenue"
    header = "Revenue"
    abbrev = "Rev."
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDSALE


class Commission(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Commission"
    header = "Commission"
    abbrev = "Commission"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.COMMISSION


class Profit(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Profit"
    header = "Profit"
    abbrev = "Profit"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDPROFIT


class SalesmanProfit(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Salesman_Profit"
    header = "Salesman Profit"
    abbrev = "Slmn Profit"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDSLMNPROFIT


class GLProfit(InvoiceDetailColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "GL_Profit"
    header = "GL Profit"
    abbrev = "GL Profit"
    metric = True
    cell_format = "InteriorCurrencyCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().c.EXTENDEDGLPROFIT
# endregion ####################################################################################################


# region CustAliasColumn  ######################################################################################
class CustAliasColumn(ColumnObject):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = None
    header = None
    abbrev = None
    metric = False
    cell_format = "InteriorTextCell"

    @staticmethod
    def apply_join(query):
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        return query.outerjoin(
            ARW_PRF_Mapping.cust_alias_01_current.__table__, and_(
                ARW_PRF_Mapping.cust_alias_01_current.__table__.c.Cust_Num ==
                ddi_data_warehouse.invoice_header.c.CUSTOMERCODE,
                ARW_PRF_Mapping.cust_alias_01_current.__table__.c.Prod_Num ==
                ddi_data_warehouse.invoice_detail.c.PRODUCTCODE
            )
        )

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def table():
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        return ARW_PRF_Mapping.cust_alias_01_current


class CustPartNum(CustAliasColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Cust_Part_Num"
    header = "Cust Part Number"
    abbrev = "Cust. Part Num."
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Cust_Part_Num


# endregion ####################################################################################################

# region ProdMainColumn  #######################################################################################
class ProdMainColumn(ColumnObject):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = None
    header = None
    abbrev = None
    metric = False
    cell_format = "InteriorTextCell"

    @staticmethod
    def apply_join(query):
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        import cep_price_console.db_management.ddi_data_warehouse as ddi_data_warehouse
        return query.outerjoin(
            ARW_PRF_Mapping.prod_main_01_current.__table__,
            ARW_PRF_Mapping.prod_main_01_current.__table__.c.Prod_Num ==
            ddi_data_warehouse.invoice_detail.c.PRODUCTCODE
        )

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def table():
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        return ARW_PRF_Mapping.prod_main_01_current


class ProdNumMain(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Prod_Num"
    header = "Product Number"
    abbrev = "Prod. Num"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Prod_Num


class DescFull(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Desc_Full"
    header = "Description"
    abbrev = "Desc."
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Desc_Full


class Status(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Status"
    header = "Product Status"
    abbrev = "Prod. Status"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Status


class ProdLine(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Prod_Line"
    header = "Product Line"
    abbrev = "Prod. Line"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Prod_Line


class PriceGroup(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Price_Group_Code"
    header = "Price Group"
    abbrev = "Price Group"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Price_Group_Code


class MfgNum(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Mfg_No"
    header = "Mfg. Num"
    abbrev = "Mfg. Num"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Mfg_No


class UPC(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "UPC"
    header = "UPC"
    abbrev = "UPC"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.UPC


class DisplayUOM(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Display_UOM"
    header = "Display UOM"
    abbrev = "Display UOM"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Display_UOM


class PriceUOM(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Price_UOM"
    header = "Price UOM"
    abbrev = "Price UOM"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Price_UOM


class PurchUOM(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Purch_UOM"
    header = "Purch UOM"
    abbrev = "Purch UOM"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Purch_UOM


class C1Cost(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "C1_Cost"
    header = "C1 Cost"
    abbrev = "C1 Cost"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.C1_Cost


class L1Price(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "L1_Price"
    header = "L1 Price"
    abbrev = "L1 Price"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.L1_Price


class PrimVend(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Primary_Vend"
    header = "Primary Vend"
    abbrev = "Prim. Vend."
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Primary_Vend


class PrimVendPartNum(ProdMainColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "Primary_Vend_Part_Num"
    header = "Prim. Vend. Part Num."
    abbrev = "Prim. Vend. Part Num."
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.Primary_Vend_Part_Num


# endregion ####################################################################################################

# region ProdUOMColumn  ########################################################################################
class ProdUOMColumn(ColumnObject):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = None
    header = None
    abbrev = None
    metric = False
    cell_format = "InteriorTextCell"

    @staticmethod
    def apply_join(query):
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        return query.outerjoin(
            ARW_PRF_Mapping.prod_uom_v2_01_current.__table__, and_(
                ARW_PRF_Mapping.prod_uom_v2_01_current.__table__.c.Prod_Num ==
                ARW_PRF_Mapping.prod_main_01_current.__table__.c.Prod_Num,
                ARW_PRF_Mapping.prod_uom_v2_01_current.__table__.c.UOM ==
                ARW_PRF_Mapping.prod_main_01_current.__table__.c.Display_UOM,
            )
        )

    @staticmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def table():
        import cep_price_console.db_management.ARW_PRF_Mapping as ARW_PRF_Mapping
        return ARW_PRF_Mapping.prod_uom_v2_01_current


class UOMFactorDesc(ProdUOMColumn):
    logger = CustomAdapter(logging.getLogger(str(__name__)), None)
    label = "UOM_Factor_Desc"
    header = "UOM Qty"
    abbrev = "UOM Qty"
    metric = False
    cell_format = "InteriorTextCell"

    @classmethod
    @debug(lvl=logging.NOTSET, prefix="")
    def column(cls):
        return cls.table().__table__.c.UOM_Factor_Desc

# endregion ####################################################################################################
